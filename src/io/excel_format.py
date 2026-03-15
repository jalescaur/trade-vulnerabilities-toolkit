"""
src/io/excel_format.py
======================
Official Excel formatting.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExcelStyle:
    title_font: Font = Font(name="Calibri", size=14, bold=True)
    header_font: Font = Font(name="Calibri", size=11, bold=True)
    body_font: Font = Font(name="Calibri", size=11, bold=False)

    title_alignment: Alignment = Alignment(horizontal="left", vertical="center")
    header_alignment: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment: Alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    header_fill: PatternFill = PatternFill("solid", fgColor="E6E6E6")
    thin_border: Border = Border(
        left=Side(style="thin", color="A0A0A0"),
        right=Side(style="thin", color="A0A0A0"),
        top=Side(style="thin", color="A0A0A0"),
        bottom=Side(style="thin", color="A0A0A0"),
    )


DEFAULT_STYLE = ExcelStyle()


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 55) -> None:
    for col in ws.columns:
        try:
            col_letter = get_column_letter(col[0].column)
        except AttributeError:
            continue

        max_len = 0
        for cell in col:
            if not cell.value:
                continue
            try:
                val = str(cell.value)
                max_len = max(max_len, len(val))
            except:
                pass
            
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def write_excel_table(
    df: pd.DataFrame,
    path: Path,
    sheet_name: str,
    title: str,
    note: str | None = None,
    percent_cols: Iterable[str] = (),
    int_cols: Iterable[str] = (),
    float_cols: Iterable[str] = (),
    currency_cols: Iterable[str] = (),
    freeze_at: str = "A3",
    style: ExcelStyle = DEFAULT_STYLE,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # FIX: Ensure ncols is at least 1 to avoid range errors
    ncols = max(1, df.shape[1])

    # 1. Title
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = style.title_font
    ws.cell(row=1, column=1).alignment = style.title_alignment
    
    # FIX: Only merge if there's more than 1 column to avoid "1 must be > 2" error
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    # 2. Header & Data
    rows = list(dataframe_to_rows(df, index=False, header=True))
    if not rows and df.empty:
        # If completely empty, just write header manually if exists
        rows = [list(df.columns)]

    for r_idx, row in enumerate(rows, start=2):
        ws.append(row)
        
        # Style Header
        if r_idx == 2:
            for c_idx in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = style.header_font
                cell.fill = style.header_fill
                cell.alignment = style.header_alignment
                cell.border = style.thin_border
        # Style Body
        else:
            for c_idx in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = style.body_font
                cell.alignment = style.body_alignment
                cell.border = style.thin_border

    ws.freeze_panes = freeze_at

    # 3. Formats
    col_index = {name: i + 1 for i, name in enumerate(df.columns)}

    def _apply_fmt(cols: Iterable[str], fmt: str):
        for col in cols:
            if col not in col_index:
                continue
            j = col_index[col]
            for i in range(3, 3 + len(df)):
                ws.cell(row=i, column=j).number_format = fmt

    _apply_fmt(percent_cols, "0.0%")
    _apply_fmt(int_cols, "#,##0")
    _apply_fmt(float_cols, "#,##0.00")
    _apply_fmt(currency_cols, "#,##0")

    # 4. Note
    if note:
        start = ws.max_row + 2
        ws.cell(row=start, column=1, value="Note:")
        ws.cell(row=start, column=1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=start, column=2, value=note)
        ws.cell(row=start, column=2).font = Font(name="Calibri", size=10)
        if ncols > 1:
            ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=ncols)

    _auto_fit_columns(ws)

    wb.save(path)
    return path